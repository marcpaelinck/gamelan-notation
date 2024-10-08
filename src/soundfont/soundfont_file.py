""" Creates an Excel soundfont definition file. This file can imported in the Viena application 
    to create a  Soundfont file (.sf2 format). See http://www.synthfont.com/The_Definitions_File.pdf

    Main method: create_soundfont_file()
"""

import os
from functools import partial

import pandas as pd
from openpyxl.styles import Font, PatternFill
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from src.common.classes import RunSettings
from src.common.constants import InstrumentType
from src.settings.settings import get_run_settings
from src.settings.settings_validation import validate_settings

MidiDict = dict[str, list[dict[str, str | int | None]]]
BOLD = Font(bold=True)


class SoundfontSheet:
    sheet: Worksheet = None
    midi_dict: MidiDict
    preset_dict: MidiDict
    settings: RunSettings

    def __init__(self, sheet: Worksheet, midi_dict: MidiDict, preset_dict: MidiDict, settings: RunSettings):
        self.sheet = sheet
        self.midi_dict = midi_dict
        self.preset_dict = preset_dict
        self.settings = settings

    def _isempty(self) -> bool:
        # openpyxl doesn't have a method for this
        return [row for row in self.sheet.rows] == []

    def _next_empty_row_nbr(self):
        return 1 if self._isempty() else self.sheet.max_row + 1

    def _add_row(self, values: list[str], bold: bool | list[bool] = False) -> None:
        # adds a new row containing the given values, each  in a separate cell.
        row = self._next_empty_row_nbr()
        for col, title in enumerate(values):
            self.sheet.cell(row=row, column=col + 1).value = title
            if (isinstance(bold, bool) and bold) or (isinstance(bold, list) and bold[col]):
                self.sheet.cell(row=row, column=col + 1).font = BOLD

    def _add_separator(self) -> None:
        # adds a separator row (containing greyed cells)
        row = self._next_empty_row_nbr()
        for col in range(1, 8):
            self.sheet.cell(row, col).fill = PatternFill(fill_type="mediumGray")

    def _add_sample_section(self) -> None:
        # Generates the first section of the Soundfont definition
        self._add_row(["Samples", "Name", "Loop S-E", "Root key", "Correction", "File name", "Folder"], bold=True)
        for records in self.midi_dict.values():
            for record in records:
                self._add_row(["", record["sample_name"], "", record["midinote"], "", record["sample"]])
        self._add_row(["Search paths:", os.path.abspath(self.settings.samples.folder)], bold=[True, False])
        self._add_separator()

    def _add_instrument_section(self) -> None:
        # Generates the second section of the Soundfont definition
        self._add_row(["Instruments", "Name", "Generators"], bold=True)
        for instrumenttype, records in self.midi_dict.items():
            self._add_row(["", instrumenttype])
            bolds = [False, False, True, False] + ([False] * len(records))
            self._add_row(["", "", "Sample name", "Global"] + [record["sample_name"] for record in records], bold=bolds)
            self._add_row(
                ["", "", "Key Range"] + [f"{record['midinote']}-{record['midinote']}" for record in records], bold=bolds
            )
            self._add_row(["", "", "Root Key"] + [record["midinote"] for record in records], bold=bolds)
        self._add_separator()

    def _add_preset_section(self) -> None:
        # Generates the third section of the Soundfont definition
        self._add_row(["Presets", "Name", "Generators", "Bank", "Preset"], bold=True)
        for (preset_name, bank, preset), records in self.preset_dict.items():
            self._add_row(["", preset_name, "", bank, preset])
            bolds = [False, False, True, False] + ([False] * len(records))
            self._add_row(
                ["", "", "Instrument name", "Global"] + [record["instrumenttype"] for record in records], bold=bolds
            )
        self._add_row(["End of data"], bold=True)
        self._add_separator()

    def create_soundfont_definition(self):
        # Generates the entire Soundfont definition content
        # This is the only method that should be called from outside this class.
        self._add_sample_section()
        self._add_instrument_section()
        self._add_preset_section()


def get_midi_dict(settings: RunSettings) -> MidiDict:
    """Reads the midinotes table and converts it into a dict of the form:
        {<instrumenttype>: [<record>, <record>, ...]}
    where <record> has the form
        {"instrumenttype": <value>, "pitch": <value>, "octave": <value>, "stroke": <value>, "midinote": <value>, "sample": <value>}

    Args:
        settings (RunSettings): used to retrieve the file path and required instrument group

    Returns:
        MidiDict:
    """
    midi_df = pd.read_csv(settings.midi.notes_filepath, sep="\t", dtype={"octave": "Int64", "midinote": "Int64"})
    to_records = partial(pd.DataFrame.to_dict, orient="records")
    midi_dict = (
        midi_df[(midi_df["instrumentgroup"] == settings.instruments.instrumentgroup) & pd.notnull(midi_df["sample"])]
        .sort_values(by=["midinote", "stroke", "instrumenttype"], ascending=[True, False, True])
        .drop(["instrumentgroup", "remark", "positions"], axis="columns")
        .drop_duplicates("midinote")
        .groupby(["instrumenttype"])[["instrumenttype", "pitch", "octave", "stroke", "midinote", "sample"]]
        .apply(to_records)
        .to_dict()
    )
    for instrumenttype, records in midi_dict.items():
        for record in records:
            record["sample_name"] = (
                instrumenttype + " " + record["pitch"] + str(record["octave"] or "") + " " + record["stroke"]
            )
    return midi_dict


def get_preset_dict(settings: RunSettings) -> MidiDict:
    """Reads the presets table and converts it into a dict of the form:
        {(<preset_name>, <bank>, <preset>): [<record>, <record>, ...]}
    where <record> has the form
        {"instrumenttype": <value>, "instrumentgroup": <value>}

    Args:
        settings (RunSettings): used to retrieve the file path and required instrument group

    Returns:
        MidiDict:
    """

    def sortkey(value: pd.Series):
        if value.name == "instrumenttype":
            return value.apply(lambda val: InstrumentType[val].sequence)
        else:
            return value

    to_records = partial(pd.DataFrame.to_dict, orient="records")
    preset_df = pd.read_csv(settings.midi.presets_filepath, sep="\t")
    preset_dict = (
        preset_df[preset_df["instrumentgroup"] == settings.instruments.instrumentgroup]
        .sort_values(by=["bank", "preset", "instrumenttype"], key=sortkey)
        .groupby(["preset_name", "bank", "preset"])[["instrumenttype", "instrumentgroup"]]
        .apply(to_records)
        .to_dict()
    )
    return preset_dict


def create_soundfont_file():
    """This method does all the work.
    All settings are read from the (YAML) settings files.
    """
    run_settings = get_run_settings()
    if run_settings.switches.validate_settings:
        validate_settings(run_settings)

    midi_dict = get_midi_dict(run_settings)
    preset_dict = get_preset_dict(run_settings)

    workbook = Workbook()
    workbook.active.title = run_settings.soundfont.sheetname
    sheet = SoundfontSheet(workbook.active, midi_dict=midi_dict, preset_dict=preset_dict, settings=run_settings)
    sheet.create_soundfont_definition()

    outfilepath = run_settings.soundfont.filepath.format(midiversion=run_settings.midi.midiversion)
    workbook.save(outfilepath)


if __name__ == "__main__":
    create_soundfont_file()
