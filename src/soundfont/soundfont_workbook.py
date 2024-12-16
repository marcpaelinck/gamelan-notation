import os

from openpyxl.styles import Font, PatternFill
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from src.common.constants import MidiDict
from src.settings.classes import RunSettings

BOLD = Font(bold=True)


class SoundfontWorkbook:
    workbook: Workbook = None
    sheet: Worksheet = None
    midi_dict: MidiDict
    preset_dict: MidiDict
    settings: RunSettings

    def __init__(self, midi_dict: MidiDict, preset_dict: MidiDict, settings: RunSettings):
        self.workbook = Workbook()
        self.sheet = self.workbook.active
        self.sheet.title = settings.soundfont.sheetname
        self.midi_dict = midi_dict
        self.preset_dict = preset_dict
        self.settings = settings

    def _isempty(self) -> bool:
        # openpyxl doesn't have a method for this
        return [row for row in self.sheet.rows] == []

    def _next_empty_row_nbr(self):
        return 1 if self._isempty() else self.sheet.max_row + 1

    def _add_row(self, values: list[str], bold: bool | list[bool] = False) -> None:
        # adds a new row containing the given values, each  in a separate cell.
        row = self._next_empty_row_nbr()
        for col, value in enumerate(values):
            self.sheet.cell(row=row, column=col + 1).value = value
            if (isinstance(bold, bool) and bold) or (isinstance(bold, list) and bold[col]):
                self.sheet.cell(row=row, column=col + 1).font = BOLD

    def _add_separator(self) -> None:
        # adds a separator row (containing greyed cells)
        row = self._next_empty_row_nbr()
        for col in range(1, 8):
            self.sheet.cell(row, col).fill = PatternFill(fill_type="mediumGray")

    def _add_sample_section(self) -> None:
        # Generates the first section of the Soundfont definition
        self._add_row(["Samples", "Name", "Loop S-E", "Root key", "Correction", "File name", "Folder"], bold=True)
        for records in self.midi_dict.values():
            for record in records:
                self._add_row(["", record["sample_name"], "", record["midinote"], "", record["sample"]])
        self._add_row(["Search paths:", os.path.abspath(self.settings.samples.folder)], bold=[True, False])
        self._add_separator()

    def _add_instrument_section(self) -> None:
        # Generates the second section of the Soundfont definition
        self._add_row(["Instruments", "Name", "Generators"], bold=True)
        for instrumenttype, records in self.midi_dict.items():
            self._add_row(["", instrumenttype])
            bolds = [False, False, True, False] + ([False] * len(records))
            self._add_row(["", "", "Sample name", "Global"] + [record["sample_name"] for record in records], bold=bolds)
            self._add_row(
                ["", "", "Key Range", ""] + [f"{record['midinote']}-{record['midinote']}" for record in records],
                bold=bolds,
            )
            self._add_row(["", "", "Root Key", ""] + [record["midinote"] for record in records], bold=bolds)
        self._add_separator()

    def _add_preset_section(self) -> None:
        # Generates the third section of the Soundfont definition
        self._add_row(["Presets", "Name", "Generators", "Bank", "Preset"], bold=True)
        for (preset_name, bank, preset), records in self.preset_dict.items():
            self._add_row(["", preset_name, "", bank, preset])
            bolds = [False, False, True, False] + ([False] * len(records))
            self._add_row(
                ["", "", "Instrument name", "Global"] + [record["instrumenttype"] for record in records], bold=bolds
            )
        self._add_row(["End of data"], bold=True)
        self._add_separator()

    def create_soundfont_definition(self):
        # Generates the entire Soundfont definition content
        # This is the only method that should be called from outside this class.
        self._add_sample_section()
        self._add_instrument_section()
        self._add_preset_section()

    def save(self):
        outfilepath = self.settings.soundfont.filepath
        self.workbook.save(outfilepath)
