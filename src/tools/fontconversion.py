"""Contains methods to convert the font of notation files from one version to the other."""

import codecs
import csv
from collections import defaultdict
from copy import copy
from glob import glob
from os import path
from pprint import pprint

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.styles import Font

from src.common.logger import Logging

logger = Logging.get_logger(__name__)

Symbol = str
Path = str
WorksheetName = str


def get_files_with_symbols(folderpath: str, symbollist: list[str]) -> dict[Path, list[Symbol]]:
    """Prints the names of Excel notation files containing one or more of the symbols
       in the symbol list.

    Args:
        folderpath (str): path to the folder containing the files to search in.
        symbollist (list[str]): list of characters to search for.
    """
    infiles = defaultdict(set)
    files = glob(path.join(folderpath, "*.xlsx"))
    wb: Workbook
    cell: Cell
    font: Font
    for file in files:
        wb = load_workbook(file)
        for sheet in [sheet for sheet in wb.worksheets if sheet.title != "formules"]:
            logger.info(path.basename(file) + " " + sheet.title)
            for row in sheet.iter_rows():
                for cell in row:
                    if (
                        cell.font.name == "Bali Music 5"
                        and cell.value
                        and any(sym in str(cell.value) for sym in symbollist)
                    ):
                        infiles[(path.basename(file), sheet.title)].add(cell.coordinate)
        wb.close()

    return infiles


def get_all_used_symbols(folderpath: str) -> tuple[set[Symbol], dict[Symbol, tuple[Path, WorksheetName]]]:
    """Prints a list of all the distinct symbols that occur in the notation files found
       in the given folder.

    Args:
        folderpath (str): path to the folder containing the files to search in.
    """
    used_symbols = set()
    files = glob(path.join(folderpath, "*.xlsx"))
    symbol_dict = dict()
    for file in files:
        wb = load_workbook(file)
        sheets = set(wb.sheetnames) - {"formules"}
        wb.close()
        for sheet in sheets:
            # logger.info(path.basename(file) + " " + sheet)
            df = pd.read_excel(file, sheet_name=sheet).iloc[:, 1:]
            uniques = set().union(
                *[
                    set(df[col][df[col].apply(lambda x: isinstance(x, str))].sum())
                    for col in df.columns
                    if set(df[col][df[col].apply(lambda x: isinstance(x, str))])
                ]
            )
            used_symbols.update(uniques)
            for s in uniques:
                symbol_dict[s] = (path.basename(file), sheet)
    return used_symbols, symbol_dict


def replace_cell_content(
    cell: Cell,
    fontreplace: tuple[str, str],
    find_replace_pairs: list[str, str],  # prio_keys: list[list[str]]
) -> dict[str, bool]:
    """Replaces individual (sequences of) characters in the given notation file and
       replaces the font in all cells.

    Args:
        worksheet (str): path to the Excel notation file.
        fontreplace (tuple[str, str]): contains source and destination font for the font conversion.
        symbolreplace (dict[list[str], list[str]], optional): _description_. Defaults to None.

    Returns:
        dict[str, bool]: values indicating whether a font change and/or a content modification was performed.
    """
    fontchanged = False
    contentchanged = False
    # Replace font
    findfont, replacefont = fontreplace
    if cell.font.name == findfont:
        oldfont: Font = cell.font
        newfont: Font = Font(
            name=replacefont,
            size=copy(oldfont.size),
            bold=copy(oldfont.bold),
            italic=copy(oldfont.italic),
            underline=copy(oldfont.underline),
            color=copy(oldfont.color),
        )
        cell.font = newfont
        fontchanged = True
    # find-replace cell content if the cell has the required font format
    if fontchanged and cell.value:
        for find, replace in find_replace_pairs:
            # convert cell content to string to avoid errors in case of numeric values
            oldval = str(copy(cell.value))
            cell.value = str(cell.value).replace(find, replace)
            contentchanged = contentchanged or (cell.value != oldval)
    return {"font": fontchanged, "content": contentchanged}


def detect_cycle(char: str, find_replace_dict: dict[str, str], passed: set[str]) -> bool:
    """Checks whether the find-replace dict contains cyclic find-replace sequences, which would give incorrect results.
       e.g. applying the find-replace dict {'A': 'BC', 'C': 'A', 'D': 'E'} to the string 'ABC' by iterating through the dict's
       items sequentially for each char in the string would yield 'BABA' instead of 'BCBA'. Applying the dict in reverse order
       would yield 'BCBBC'.
       This method is applied recursively to each 'priority key' which is a set of keys that contain dict keys in their
       'replace' string, which in turn might cause a cycle. A recursion ends either if a cycle is detected or if a char does not
       occur in the dict.
       In our example, ['A', 'C'] would be the set of priority keys. Applying the method recursively  to key 'A' would yield
       the following (char, target, passed) values:
       ('A', 'BC',  passed=['A'])-> ('B', None, passed=['A']) -> return False (end of recursion)
                                '-> ('C', 'A', passed=['A']) -> return True (end of recursion)
    Args:
        key (str): priority key to check.
        find_replace_dict (dict[str, str]): find-replace dict.
        encountered (list[str], optional): list that keeps track of the keys that were already passed during a recursion.
          A cycle is detected if the same key is passed twice.  Defaults to list().

    Returns:
        bool: _description_
    """
    if char in passed:
        # cycle detected: end of recursion
        return True

    target = find_replace_dict.get(char, None)
    if target:
        # start new recursions
        passed.add(char)
        return any(detect_cycle(symbol, find_replace_dict, passed) for symbol in target)
    # char does not occur in the dict: end of recursion
    return False


def sort_prio_keys(keys: list[str], symdict: dict[str, str]) -> list[str]:
    """Sorts the keys that must be converted with priority in such a way that each key K
       is followed by keys that are not contained in K's replacement string. This will
       avoid the situation that the replacement string of one key is (partly) replaced
       again by a subsequent key.
       Note: For such a sort order to be possible we will first need to check that there
       are not circular replacement cycles (e.g. key1 replaces key2 and key2 replaces key1).
       See method `detect_cycle`.

    Args:
        keys (list[str]): The list of
        symdict (dict[str, str]): _description_

    Returns:
        list[str]: _description_
    """
    sortedkeys = []
    for key in keys:
        after = [k for k, v in symdict.items() if key in v]
        found = next((k for k in sortedkeys if k in after), None)
        if found:
            sortedkeys.insert(sortedkeys.index(found), key)
        else:
            sortedkeys.append(key)
    return sortedkeys


def create_find_replace_pairs(symbolreplacetablepath: str) -> list[list[str, str]]:
    sym_dict = pd.read_csv(symbolreplacetablepath, sep="\t", quoting=csv.QUOTE_NONE).to_dict(orient="split")
    symbolreplacedict = {source: target for source, target in sym_dict["data"]}

    # Determine which symbols need to be converted first. These are symbols that occur in one of the target strings.
    # By treating these symbols first, we can avoid "follow-up" replacements (replacing an already replaced string).
    # We raise an exception if there is a circular reference within these keys, because this situation will always
    # cause incorrect replacements.
    prio_keys = [key for key in symbolreplacedict.keys() if any(key in val for val in symbolreplacedict.values())]
    if any(detect_cycle(symbol, symbolreplacedict, set()) for symbol in prio_keys):
        raise ValueError(
            f"The keys {prio_keys} contain a cyclic find-replace sequence, which might cause incorrect results."
        )
    prio_keys = sort_prio_keys(prio_keys, symbolreplacedict)
    # Generate a list of find-replace pairs in the required sequence (prio keys first)
    find_replace_pairs = [[find, symbolreplacedict[find]] for find in prio_keys] + [
        [find, replace] for find, replace in symbolreplacedict.items() if find not in prio_keys
    ]
    return find_replace_pairs


def substitutefont_xlsx_file(
    filepath: str,
    savepath: str,
    fontreplacepair: list[str, str],
    find_replace_list: list[list[str, str]] = None,
    symbolreplacetablepath: str = None,
):
    """Substitutes the music font in all Excel notation files found in the given folder.
       This method also replaces symbols in necessary.

    Args:
        filepath (str): path to the notation file.
        savepath (str): path to the folder where converted files should be saved.
        fontfrom (str): font name to replace.
        fontto (str): name of the new font.
        symbolreplacetablepath (str): path to the conversion file, which indicates how (groups of) symbols should be converted.
    """
    find_replace_pairs = find_replace_list or create_find_replace_pairs(symbolreplacetablepath)
    if not find_replace_pairs:
        raise ValueError(
            "Missing find-replace data. Either parameter 'find_replace_list' or 'symbolreplacetablepath' should have a value other than `None`."
        )

    changes = {"font": False, "content": False}
    wb = load_workbook(filepath)
    for sheet in wb.worksheets:
        if sheet.title != "formules":
            for row in sheet.rows:
                for cell in row:
                    cell_changes = replace_cell_content(
                        cell=cell,
                        fontreplace=fontreplacepair,
                        find_replace_pairs=find_replace_pairs,
                    )
                    changes = {key: (changes[key] or cell_changes[key]) for key in changes.keys()}

    changed_items = " and ".join([key for key, modified in changes.items() if modified] or ["nothing"])
    logger.info(f"{path.basename(filepath)}: {changed_items} changed")
    wb.save(path.join(savepath, path.basename(filepath).replace("_4", "_5")))


def substitutefont_tsv_file(
    filepath: str,
    savepath: str,
    find_replace_list: list[list[str, str]] = None,
    symbolreplacetablepath: str = None,
):
    """Substitutes the music font in a single tsv notation file.
       This method also replaces symbols in necessary.

    Args:
        filepath (str): path to the tsv source file.
        savepath (str): path for the converted file.
        find_replace_list (list[list[str, str]]): list of find-replace pairs. Either this value or symbolreplacetablepath should be given.
        symbolreplacetablepath (str): path to the conversion file, which indicates how (groups of) symbols should be converted.
    """
    find_replace_pairs = find_replace_list or create_find_replace_pairs(symbolreplacetablepath)
    if not find_replace_pairs:
        raise ValueError(
            "Missing find-replace data. Either parameter 'find_replace_list' or 'symbolreplacetablepath' should have a value other than `None`."
        )
    with codecs.open(filepath, encoding="UTF-8", mode="r") as infile:
        inlines = infile.readlines()

    outlines = []
    for inline in inlines:
        outline = copy(inline)
        # Each line starts with a tag followed by a tab character and the notation.
        # The tag should not be converted.
        # Comment and metadata lines should be left unchanged.
        tag = outline.split("\t")[0]
        if not tag in ["comment", "metadata"]:
            notation = outline[len(tag) :]
            for find, replace in find_replace_pairs:
                notation = notation.replace(find, replace)
            outline = tag + notation
        outlines.append(outline)
    with codecs.open(savepath, encoding="UTF-8", mode="w") as outfile:
        outfile.writelines(outlines)


def substitutefont_all_xlsx_files(
    folderpath: str, savepath: str, fontreplacepair: list[str, str], symbolreplacetablepath: str
):
    """Substitutes the music font in all Excel notation files found in the given folder.
       This method also replaces symbols in necessary.

    Args:
        folderpath (str): path to the folder containing the files to search in.
        savepath (str): path to the folder where converted files should be saved.
        fontreplacepair (list[str, str]): names of the font to replace and of the new font.
        symbolreplacetablepath (str): path to the conversion file, which indicates how (groups of) symbols should be converted.
    """
    find_replace_pairs = create_find_replace_pairs(symbolreplacetablepath)

    xls_files = glob(path.join(folderpath, "*.xlsx"))
    for filepath in xls_files:
        substitutefont_xlsx_file(filepath, savepath, fontreplacepair, find_replace_pairs)


SYMBOLREPLACETABLEPATH = "./config/convert4to5font.tsv"

XL_FOLDERPATH = "C:/Users/marcp/Documents/administratie/_VRIJETIJD_REIZEN/Gamelangroepen/Studiemateriaal/Muzieknotatie"
XL_SAVEFOLDERPATH = (
    "C:/Users/marcp/Documents/administratie/_VRIJETIJD_REIZEN/Gamelangroepen/Studiemateriaal/Muzieknotatie/balimusic5"
)
if __name__ == "__main__":
    # substitutefont_tsv_file(
    #     filepath=os.path.join(TSV_SETTING.datapath, TSV_SETTING.infilename),
    #     savepath=os.path.join(TSV_SETTING.datapath, "Margapati reyong_font5.tsv"),
    #     symbolreplacetablepath=SYMBOLREPLACETABLEPATH,
    # )
    result = get_files_with_symbols(XL_FOLDERPATH, ["1", "2", "3", "5", "6", "7"])
    pprint(result)
