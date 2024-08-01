import json
import os
from collections import defaultdict
from glob import glob
from os import path
from pprint import pprint

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font
from scipy.io import wavfile

from notation_classes import InstrumentTag
from notation_constants import (
    InstrumentGroup,
    InstrumentPosition,
    InstrumentType,
    MutingType,
    NoteType,
)
from src.utils import create_instrumentrange_lookup, create_note_to_midi_lookup


def get_all_tags():
    folder = "G:\\Marc\\documents-backup-2jun24-08h00\\Documents\\administratie\\_VRIJETIJD_REIZEN\\Gamelangroepen\\Studiemateriaal\\Muzieknotatie"
    files = [file for file in glob(path.join(folder, "*.xlsx")) if not path.basename(file).startswith("_")]
    tagdict = dict()
    for file in files:
        taglist = set(pd.read_excel(file).iloc[:, 0].unique())
        tagdict.update({tag: path.basename(file) for tag in taglist if isinstance(tag, str)})
    pprint(sorted(list(tagdict.keys())))
    pprint(tagdict)


def map_positions():
    instrumenttag_dict = pd.read_csv("./settings/instrumenttags_1.csv", sep="\t").to_dict(orient="records")
    tag0 = InstrumentTag.model_validate(instrumenttag_dict[0])
    tags = [InstrumentTag.model_validate(tag) for tag in instrumenttag_dict]
    mapping_r = [
        (
            ["reyong13", "reyong1+3", "reyong 1/3", "reyong p"],
            [InstrumentPosition.REYONG_1, InstrumentPosition.REYONG_3],
        ),
        (
            ["reyong24", "reyong2+4", "reyong 2.4", "reyong s", "reyong 2/4"],
            [InstrumentPosition.REYONG_2, InstrumentPosition.REYONG_4],
        ),
        (["reyong12", "reyong1+2"], [InstrumentPosition.REYONG_1, InstrumentPosition.REYONG_2]),
        (["reyong34", "reyong3+4"], [InstrumentPosition.REYONG_3, InstrumentPosition.REYONG_4]),
        (["reyong1-3"], [InstrumentPosition.REYONG_1, InstrumentPosition.REYONG_3]),
        (["reyong2-4"], [InstrumentPosition.REYONG_2, InstrumentPosition.REYONG_4]),
        (
            ["reyong1-4"],
            [
                InstrumentPosition.REYONG_1,
                InstrumentPosition.REYONG_2,
                InstrumentPosition.REYONG_3,
                InstrumentPosition.REYONG_4,
            ],
        ),
        (["reyong1", "reyong(1)", "reyong 1"], [InstrumentPosition.REYONG_1]),
        (["reyong2", "reyong(2)", "reyong 2"], [InstrumentPosition.REYONG_2]),
        (["reyong3", "reyong(3)", "reyong 3"], [InstrumentPosition.REYONG_3]),
        (["reyong4", "reyong(4)", "reyong 4"], [InstrumentPosition.REYONG_4]),
        (
            ["Reyong", "reyong", "rey/", "/rey", "+rey"],
            [
                InstrumentPosition.REYONG_1,
                InstrumentPosition.REYONG_2,
                InstrumentPosition.REYONG_3,
                InstrumentPosition.REYONG_4,
            ],
        ),
    ]
    mapping_g = [
        (
            ["gangsa p/s"],
            [
                InstrumentPosition.PEMADE_POLOS,
                InstrumentPosition.PEMADE_SANGSIH,
                InstrumentPosition.KANTILAN_POLOS,
                InstrumentPosition.KANTILAN_SANGSIH,
            ],
        ),
        (["gangsa p", "gang p", "ga p", "/ga.p"], [InstrumentPosition.PEMADE_POLOS, InstrumentPosition.KANTILAN_POLOS]),
        (["gangsa s", "ga s"], [InstrumentPosition.PEMADE_SANGSIH, InstrumentPosition.KANTILAN_SANGSIH]),
        (["pemade p"], [InstrumentPosition.PEMADE_POLOS]),
        (["pemade s"], [InstrumentPosition.PEMADE_SANGSIH]),
        (["kantilan p"], [InstrumentPosition.KANTILAN_POLOS]),
        (["kantilan s"], [InstrumentPosition.KANTILAN_SANGSIH]),
        (["pemade"], [InstrumentPosition.PEMADE_POLOS, InstrumentPosition.PEMADE_SANGSIH]),
        (["kantilan"], [InstrumentPosition.KANTILAN_POLOS, InstrumentPosition.KANTILAN_SANGSIH]),
        (
            ["ga ", "ga+", "ga/", "/ga", "/gang", "gang.", "ga4", "gangs4", "gangsa", "(ga)", "/ ga"],
            [
                InstrumentPosition.PEMADE_POLOS,
                InstrumentPosition.PEMADE_SANGSIH,
                InstrumentPosition.KANTILAN_POLOS,
                InstrumentPosition.KANTILAN_SANGSIH,
            ],
        ),
    ]
    for mytag in tags:
        mytag.positions = [
            InstrumentPosition[instr.value] for instr in mytag.instruments if instr.value in InstrumentPosition
        ]
        for values, positions in mapping_r:
            if any(val in mytag.tag for val in values):
                mytag.positions.extend(positions)
                break
        for values, positions in mapping_g:
            if any(val in mytag.tag for val in values):
                mytag.positions.extend(positions)
                break
    tags_dict = [mytag.model_dump() for mytag in tags]
    print([t.tag for t in tags if not t.positions])
    tags_df = pd.DataFrame.from_records(tags_dict)
    tags_df.sort_values(by="tag", key=lambda col: col.str.lower())
    tags_df.to_csv("./settings/instrumenttags.csv", sep="\t", index=False)


def merge_parts(datapath: str, basefile: str, mergefile: str, resultfile: str):
    columns = ["tag"] + ["BEAT" + str(i) for i in range(1, 33)]
    sortingorder = [
        "comment",
        "metadata",
        "ugal",
        "gangsa p",
        "gangsa s",
        "reyong1-4",
        "reyong1",
        "reyong2",
        "reyong3",
        "reyong4",
        "calung",
        "jegog",
        "gong",
        "kendang",
    ]
    # Load the scores in dataframes
    base_df = pd.read_csv(
        path.join(datapath, basefile), sep="\t", names=columns, skip_blank_lines=False, encoding="UTF-8"
    )
    merge_df = pd.read_csv(
        path.join(datapath, mergefile), sep="\t", names=columns, skip_blank_lines=False, encoding="UTF-8"
    )
    # Drop all empty columns
    base_df.dropna(how="all", axis="columns", inplace=True)
    merge_df.dropna(how="all", axis="columns", inplace=True)
    # Number the systems: blank lines denote start of new system. Then delete blank lines.
    base_df["sysnr"] = base_df["tag"].isna().cumsum()[~base_df["tag"].isna()] + 1
    merge_df["sysnr"] = merge_df["tag"].isna().cumsum()[~merge_df["tag"].isna()] + 1
    merge_df = merge_df[~merge_df["tag"].isin(["gangsa p", "gangsa s"])]
    # Drop all empty rows
    base_df.dropna(how="all", axis="rows", inplace=True)
    merge_df.dropna(how="all", axis="rows", inplace=True)
    # Concatenate both tables
    new_df = pd.concat([base_df, merge_df], ignore_index=True)
    # Sort the new table
    new_df["tagid"] = new_df["tag"].apply(lambda tag: sortingorder.index(tag))
    new_df.sort_values(by=["sysnr", "tagid"], inplace=True, ignore_index=True)
    # Add empty lines between systems
    mask = new_df["sysnr"].ne(new_df["sysnr"].shift(-1))
    empties = pd.DataFrame("", index=mask.index[mask] + 0.5, columns=new_df.columns)
    new_df = pd.concat([new_df, empties]).sort_index().reset_index(drop=True).iloc[:-1]
    # Drop sysnr and tagid columns
    new_df.drop(["sysnr", "tagid"], axis="columns", inplace=True)
    new_df.to_csv(path.join(datapath, resultfile), sep="\t", index=False, header=False)


def rename_files(folderpath: str, filter: str, findtext: str, replacetext: str, print_only: bool = True):
    filelist = glob(path.join(folderpath, filter))
    if isinstance(findtext, str) and isinstance(replacetext, str):
        findtext = [findtext]
        replacetext = [replacetext]
    for filename in filelist:
        newfilename = filename
        for pos, find in enumerate(findtext):
            newfilename = newfilename.replace(find, replacetext[pos])
            if newfilename != filename:
                break
        if print_only:
            print(f"{os.path.basename(filename)} -> {os.path.basename(newfilename)}")
        else:
            os.rename(filename, newfilename)
    if print_only:
        print("No files were renamed (print_only==True)")


def rename_notes_in_filenames(folderpath: str, group: InstrumentGroup, print_only: bool = True):
    # In the filenames of Bali Gamelan Samples, Ding 1 is always the first note in an instrument's range.
    # This function renames the files according to the (relative) naming in this application.
    instrdict = {
        "Kantil": InstrumentType.KANTILAN,
        "Pemade": InstrumentType.PEMADE,
        "Rejong": InstrumentType.REYONG,
        "Ugal": InstrumentType.UGAL,
    }
    filenotes = sum([[f"Ding {i}", f"Dong {i}", f"Deng {i}", f"Dung {i}", f"Dang {i}"] for i in range(1, 4)], [])
    lookup = create_instrumentrange_lookup(group)
    lookup = {
        instrtype: [
            note
            for note in notes
            if note.isnote and note.note.type == NoteType.MELODIC and note.mutingtype == MutingType.OPEN
        ]
        for instrtype, notes in lookup.items()
        if instrtype in instrdict.values()
    }
    for instr_name, instr_type in instrdict.items():
        notedict = {filenotes[lookup[instr_type].index(note)]: note.value for note in lookup[instr_type]}
        rename_files(
            folderpath,
            filter=f"*{instr_name}*.wav",
            findtext=list(notedict.keys()),
            replacetext=list(notedict.values()),
            print_only=print_only,
        )


def convert_to_wav_pcm(filepath_in: str, filepath_out: str):
    rate, data = wavfile.read(filepath_in)
    data = data.astype("int16")
    wavfile.write(filepath_out, rate, data)


def get_files_with_symbols():
    folderpath = "G:\\Marc\\documents-backup-2jun24-08h00\\Documents\\administratie\\_VRIJETIJD_REIZEN\\Gamelangroepen\\Studiemateriaal\\Muzieknotatie"
    # search = {"ê", "ë", "Ë", "â", "ä", "ï", "Ï", "ĭ", "ö", "û", "Ü", "ü"}
    search = ["¯", "«", "©", "¬", "®", "°"]  # onjuiste kendang
    # search = ["ṁ", "ṃ", "ṅ", "ṇ"] # correcte kendang-noten (ook nog m en n) - nergens gebruikt

    infiles = defaultdict(set)
    files = glob(path.join(folderpath, "*.xlsx"))
    for file in files:
        wb = load_workbook(file)
        sheets = set(wb.sheetnames) - {"formules"}
        wb.close()
        for sheet in sheets:
            print(os.path.basename(file) + " " + sheet)
            df = pd.read_excel(file, sheet_name=sheet).iloc[:, 1:]
            uniques = set().union(
                *[
                    set(df[col][df[col].apply(lambda x: isinstance(x, str))].sum())
                    for col in df.columns
                    if set(df[col][df[col].apply(lambda x: isinstance(x, str))])
                ]
            )
            for sym in search:
                if sym in uniques:
                    infiles[sym].add(os.path.basename(file))

    pprint(infiles)


def get_all_used_symbols():
    all_uniques = set()
    folderpath = "G:\\Marc\\documents-backup-2jun24-08h00\\Documents\\administratie\\_VRIJETIJD_REIZEN\\Gamelangroepen\\Studiemateriaal\\Muzieknotatie"
    files = glob(path.join(folderpath, "*.xlsx"))
    where = dict()
    for file in files:
        wb = load_workbook(file)
        sheets = set(wb.sheetnames) - {"formules"}
        wb.close()
        for sheet in sheets:
            print(os.path.basename(file) + " " + sheet)
            df = pd.read_excel(file, sheet_name=sheet).iloc[:, 1:]
            uniques = set().union(
                *[
                    set(df[col][df[col].apply(lambda x: isinstance(x, str))].sum())
                    for col in df.columns
                    if set(df[col][df[col].apply(lambda x: isinstance(x, str))])
                ]
            )
            all_uniques.update(uniques)
            for s in uniques:
                where[s] = (os.path.basename(file), sheet)
    print(all_uniques)
    pprint(where)


def replace_cell_content(fromfont: str = None, tofont: str = None, search: str = None, replace: str = None):
    folderpath = "G:\\Marc\\documents-backup-2jun24-08h00\\Documents\\administratie\\_VRIJETIJD_REIZEN\\Gamelangroepen\\Studiemateriaal\\Muzieknotatie\\balimusic4"
    savepath = "G:\\Marc\\documents-backup-2jun24-08h00\\Documents\\administratie\\_VRIJETIJD_REIZEN\\Gamelangroepen\\Studiemateriaal\\Muzieknotatie"
    fromfont = None  # "Bali Music 3"
    tofont = None  # "Bali Music 4"
    files = [
        "Hujan Mas.xlsx",
        "Jauk keras.xlsx",
        "Kebyar Amstelveen.xlsx",
        "Manuk Rawa.xlsx",
        "Sidakarya.xlsx",
        "Sidakarya1.xlsx",
    ]
    searchreplace = [("¯", "ṃ"), ("«", "ṇ"), ("©", "n"), ("¬", "ṅ"), ("®", "m"), ("°", "ṁ")]
    files = [path.join(folderpath, file) for file in files] or glob(path.join(folderpath, "*.xlsx"))
    fontchanged = False
    contentchanged = False
    for file in files:
        wb = load_workbook(file)
        for sheet in wb.worksheets:
            for row in sheet.rows:
                for cell in row:
                    # Replace font
                    if fromfont and tofont:
                        if cell.font.name == fromfont:
                            oldfont: Font = cell.font
                            newfont: Font = Font(
                                name=tofont,
                                size=oldfont.size,
                                bold=oldfont.bold,
                                italic=oldfont.italic,
                                underline=oldfont.underline,
                                color=oldfont.color,
                            )
                            cell.font = newfont
                            fontchanged = True
                    # find-replace in content
                    if searchreplace and isinstance(cell.value, str):
                        for find, replace in searchreplace:
                            oldval = cell.value
                            cell.value = cell.value.replace(find, replace)
                            contentchanged = contentchanged or (cell.value != oldval)
        print(
            f"{os.path.basename(file)}: {'font changed' if fontchanged else ''}{'and' if fontchanged and contentchanged else ''}{'value(s) changed' if contentchanged else ''}"
        )
        wb.save(os.path.join(savepath, os.path.basename(file)))


if __name__ == "__main__":
    replace_cell_content()
    # foldername = "G:/Marc/documents-backup-2jun24-08h00/Documents/administratie/_VRIJETIJD_REIZEN/Gamelangroepen/Studiemateriaal/audio-samples/GONG KEBYAR"
    # foldername_out = "G:/Marc/documents-backup-2jun24-08h00/Documents/administratie/_VRIJETIJD_REIZEN/Gamelangroepen/Studiemateriaal/audio-samples/GONG KEBYAR WAV"
    # filename = "02 Kantil Ombak DONG0.wav"
    # convert_to_wav_pcm(os.path.join(foldername, filename), os.path.join(foldername_out, filename))